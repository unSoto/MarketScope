import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import pandas as pd
from typing import List, Dict, Any
import logging
from PIL import Image, ImageTk
import os

from database.db_manager import DatabaseManager
from parser.hh_parser import HHParser

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Color scheme
COLORS = {
    'primary': '#5E01A7',      # Purple
    'secondary': '#F9432E',    # Orange/Red
    'accent': '#C1009A',       # Pink/Magenta
    'background': '#F8F9FA',   # Light gray
    'surface': '#FFFFFF',      # White
    'text': '#2C3E50',         # Dark gray
    'text_secondary': '#6C757D', # Light gray
    'success': '#28A745',      # Green
    'warning': '#FFC107',      # Yellow
    'error': '#DC3545'         # Red
}


class VacancyApp:
    """Main GUI application for vacancy scraping and management."""

    def __init__(self, root):
        """Initialize the main application window."""
        self.root = root
        self.root.title("MarketScope - HH.ru Vacancy Scraper")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)

        # Configure root window styling
        self.root.configure(bg=COLORS['background'])

        # Initialize components
        self.db_manager = DatabaseManager()
        self.parser = HHParser()

        # Search parameters
        self.keyword_var = tk.StringVar()
        self.location_var = tk.StringVar()
        self.experience_var = tk.StringVar(value="")
        self.current_search_results = []  # Store current search results

        # Create GUI elements
        self._create_header()
        self._create_menu_bar()
        self._create_search_frame()
        self._create_results_frame()
        self._create_analytics_section()
        self._create_status_bar()

        # Load initial data
        self._load_vacancies()

        # Show main interface
        self._show_main_interface()

    def _show_main_interface(self):
        """Show the main search and results interface."""
        # Show search frame
        self.search_frame.pack(fill="x", padx=10, pady=(10, 0))

        # Show results frame
        self.results_frame.pack(fill="both", expand=True, padx=10, pady=(10, 0))

        # Show analytics section (initially hidden, but button visible)
        self.analytics_frame.pack(fill="x", padx=10, pady=(10, 0))

    def _create_navigation(self):
        """Create navigation buttons."""
        nav_frame = tk.Frame(self.root, bg=COLORS['surface'], relief="groove", bd=2)
        nav_frame.pack(fill="x", padx=10, pady=(10, 0))

        # Navigation buttons
        self.search_nav_button = tk.Button(
            nav_frame,
            text="🔍 Поиск вакансий",
            command=self._show_search_page,
            bg=COLORS['primary'],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        self.search_nav_button.pack(side="left", padx=(10, 5), pady=10)

        self.analytics_nav_button = tk.Button(
            nav_frame,
            text="📊 Аналитика",
            command=self._show_analytics_page,
            bg=COLORS['surface'],
            fg=COLORS['primary'],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        self.analytics_nav_button.pack(side="left", padx=(5, 10), pady=10)

    def _create_main_container(self):
        """Create main container for pages."""
        self.main_container = tk.Frame(self.root, bg=COLORS['background'])
        self.main_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def _show_search_page(self):
        """Show search page."""
        # Hide analytics frame
        if hasattr(self, 'analytics_frame'):
            self.analytics_frame.pack_forget()

        # Show search and results frames
        self.search_frame.pack(fill="x", padx=0, pady=(10, 0))
        self.results_frame.pack(fill="both", expand=True, padx=0, pady=(10, 0))

        # Update button styles
        self.search_nav_button.config(bg=COLORS['primary'], fg="white")
        self.analytics_nav_button.config(bg=COLORS['surface'], fg=COLORS['primary'])

        # Load vacancies when returning to search page
        self._load_vacancies()

        # Update status bar
        vacancies = self.db_manager.get_all_vacancies()
        self.status_var.set(f"Загружено вакансий: {len(vacancies)}")

        # Make sure search controls are enabled
        self._set_search_state(True)

    def _show_analytics_page(self):
        """Show analytics page."""
        # Hide search and results frames
        self.search_frame.pack_forget()
        self.results_frame.pack_forget()

        # Show analytics frame
        self.analytics_frame.pack(fill="both", expand=True, padx=0, pady=0)

        # Update button styles
        self.search_nav_button.config(bg=COLORS['surface'], fg=COLORS['primary'])
        self.analytics_nav_button.config(bg=COLORS['primary'], fg="white")

        # Update analytics data
        self._update_analytics()

    def _create_header(self):
        """Create the application header with logo."""
        header_frame = tk.Frame(self.root, bg=COLORS['primary'], height=80)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)

        # Try to load logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                logo_image = logo_image.resize((60, 60), Image.Resampling.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_image)

                logo_label = tk.Label(header_frame, image=self.logo_photo, bg=COLORS['primary'])
                logo_label.pack(side="left", padx=20, pady=10)
        except Exception as e:
            logger.warning(f"Could not load logo: {e}")

        # Title
        title_label = tk.Label(
            header_frame,
            text="MarketScope",
            font=("Arial", 24, "bold"),
            fg="white",
            bg=COLORS['primary']
        )
        title_label.pack(side="left", padx=20, pady=10)

        # Subtitle
        subtitle_label = tk.Label(
            header_frame,
            text="HH.ru Vacancy Scraper",
            font=("Arial", 12),
            fg="white",
            bg=COLORS['primary']
        )
        subtitle_label.pack(side="left", padx=0, pady=10)

    def _create_menu_bar(self):
        """Create the menu bar."""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Экспорт в Excel", command=self._export_to_excel)
        file_menu.add_command(label="Экспорт в CSV", command=self._export_to_csv)
        file_menu.add_separator()
        file_menu.add_command(label="Очистить базу данных", command=self._clear_database)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.root.quit)

        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Вид", menu=view_menu)
        view_menu.add_command(label="Обновить данные", command=self._load_vacancies)
        view_menu.add_command(label="Статистика", command=self._show_statistics)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Справка", menu=help_menu)
        help_menu.add_command(label="О программе", command=self._show_about)

    def _create_search_frame(self):
        """Create the search parameters frame."""
        self.search_frame = tk.Frame(self.root, bg=COLORS['surface'], relief="groove", bd=2)

        # Title
        title_label = tk.Label(
            self.search_frame,
            text="🔍 Параметры поиска",
            font=("Arial", 14, "bold"),
            bg=COLORS['surface'],
            fg=COLORS['primary']
        )
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 15), sticky="w", padx=10)

        # Configure grid weights for proper resizing
        self.search_frame.grid_columnconfigure(1, weight=1)
        self.search_frame.grid_columnconfigure(3, weight=1)

        # Keyword input
        keyword_label = tk.Label(self.search_frame, text="Ключевое слово:", bg=COLORS['surface'], fg=COLORS['text'])
        keyword_label.grid(row=1, column=0, sticky="w", padx=10, pady=5)
        keyword_entry = ttk.Entry(self.search_frame, textvariable=self.keyword_var, width=30, font=("Arial", 10))
        keyword_entry.grid(row=1, column=1, sticky="ew", padx=10, pady=5)

        # Location selection
        location_label = tk.Label(self.search_frame, text="Местоположение:", bg=COLORS['surface'], fg=COLORS['text'])
        location_label.grid(row=1, column=2, sticky="w", padx=10, pady=5)
        location_combo = ttk.Combobox(self.search_frame, textvariable=self.location_var, width=28, font=("Arial", 10), state="readonly")
        location_combo['values'] = [
            '', 'Москва', 'Санкт-Петербург', 'Екатеринбург', 'Новосибирск',
            'Краснодар', 'Нижний Новгород', 'Казань', 'Челябинск', 'Омск',
            'Самара', 'Ростов-на-Дону', 'Уфа', 'Красноярск', 'Воронеж',
            'Волгоград', 'Пермь'
        ]
        location_combo.grid(row=1, column=3, sticky="ew", padx=10, pady=5)

        # Experience selection
        experience_label = tk.Label(self.search_frame, text="Опыт работы:", bg=COLORS['surface'], fg=COLORS['text'])
        experience_label.grid(row=2, column=0, sticky="w", padx=10, pady=5)
        experience_combo = ttk.Combobox(self.search_frame, textvariable=self.experience_var, width=28, font=("Arial", 10), state="readonly")
        experience_combo['values'] = [
            '', 'Нет опыта', 'От 1 до 3 лет', 'От 3 до 6 лет', 'Более 6 лет'
        ]
        experience_combo.grid(row=2, column=1, sticky="ew", padx=10, pady=5)

        # Buttons frame
        buttons_frame = tk.Frame(self.search_frame, bg=COLORS['surface'])
        buttons_frame.grid(row=3, column=0, columnspan=4, pady=15)

        # Search button
        search_button = tk.Button(
            buttons_frame,
            text="🚀 Начать поиск",
            command=self._start_search,
            bg=COLORS['secondary'],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        search_button.pack(side="left", padx=(0, 10))

        # Clear database button
        clear_button = tk.Button(
            buttons_frame,
            text="🗑️ Очистить всё",
            command=self._clear_database,
            bg=COLORS['accent'],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        clear_button.pack(side="left")

        # Progress frame
        progress_frame = tk.Frame(self.search_frame, bg=COLORS['surface'])
        progress_frame.grid(row=4, column=0, columnspan=4, sticky="ew", padx=10)

        # Progress label
        self.progress_label = tk.Label(
            progress_frame,
            text="Готов к работе",
            bg=COLORS['surface'],
            fg=COLORS['text_secondary'],
            font=("Arial", 9)
        )
        self.progress_label.pack(anchor="w", pady=(0, 5))

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            style="TProgressbar"
        )
        self.progress_bar.pack(fill="x", pady=(0, 5))

        # Status text
        self.status_text = tk.Label(
            progress_frame,
            text="",
            bg=COLORS['surface'],
            fg=COLORS['text_secondary'],
            font=("Arial", 8)
        )
        self.status_text.pack(anchor="w")

    def _create_results_frame(self):
        """Create the results display frame."""
        self.results_frame = tk.Frame(self.root, bg=COLORS['surface'], relief="groove", bd=2)

        # Title with icon
        title_label = tk.Label(
            self.results_frame,
            text="📋 Результаты поиска",
            font=("Arial", 14, "bold"),
            bg=COLORS['surface'],
            fg=COLORS['primary']
        )
        title_label.pack(anchor="w", padx=10, pady=(10, 5))

        # Create modern styled frame for treeview
        tree_frame = tk.Frame(self.results_frame, bg="white", relief="flat")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Style the treeview
        style = ttk.Style()

        # Configure treeview item styling
        style.configure("Modern.Treeview",
                       background="white",
                       foreground="#2C3E50",  # Dark gray text
                       rowheight=35,
                       fieldbackground="white",
                       borderwidth=0,
                       font=("Arial", 10))

        # Configure treeview heading styling
        style.configure("Modern.Treeview.Heading",
                       background=COLORS['primary'],
                       foreground="black",
                       font=("Arial", 11, "bold"),
                       relief="flat")

        # Configure treeview item states
        style.map("Modern.Treeview.Heading",
                 background=[("active", COLORS['secondary'])])

        style.map("Modern.Treeview",
                 background=[("selected", COLORS['background'])],
                 foreground=[("selected", COLORS['primary'])])

        # Force apply the heading style
        style.configure("Treeview.Heading",
                       background=COLORS['primary'],
                       foreground="white",
                       font=("Arial", 11, "bold"),
                       relief="flat")

        style.map("Treeview.Heading",
                 background=[("active", COLORS['secondary'])])

        # Additional styling for better visibility
        style.configure("Treeview",
                       background="white",
                       foreground="#2C3E50",
                       fieldbackground="white")

        style.configure("Treeview.Heading",
                       background=COLORS['primary'],
                       foreground="white")

        # Create treeview for results - only title and company
        columns = ('title', 'company')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                               height=20, style="Modern.Treeview")

        # Define column headings
        self.tree.heading('title', text='📋 Название вакансии')
        self.tree.heading('company', text='🏢 Компания')

        # Configure column widths
        self.tree.column('title', width=450, minwidth=300)
        self.tree.column('company', width=250, minwidth=150)

        # Add only vertical scrollbar (no horizontal scrollbar)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        # Pack the treeview and vertical scrollbar
        self.tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

        # Add context menu
        self._create_context_menu()

    def _create_analytics_section(self):
        """Create analytics section below results."""
        # Create analytics frame that will be packed below results
        self.analytics_frame = tk.Frame(self.root, bg=COLORS['surface'], relief="groove", bd=2)

        # Title with toggle button
        title_frame = tk.Frame(self.analytics_frame, bg=COLORS['surface'])
        title_frame.pack(fill="x", padx=10, pady=(10, 5))

        # Analytics toggle button
        self.analytics_visible = tk.BooleanVar(value=False)

        def toggle_analytics():
            if self.analytics_visible.get():
                self.analytics_content.pack_forget()
                self.toggle_button.config(text="📊 Показать аналитику")
                self.analytics_visible.set(False)
            else:
                self.analytics_content.pack(fill="both", expand=True, padx=10, pady=(0, 10))
                self.toggle_button.config(text="📊 Скрыть аналитику")
                self.analytics_visible.set(True)
                # Update analytics when showing
                self._update_analytics()

        self.toggle_button = tk.Button(
            title_frame,
            text="📊 Показать аналитику",
            command=toggle_analytics,
            bg=COLORS['primary'],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=15,
            pady=5,
            cursor="hand2"
        )
        self.toggle_button.pack(anchor="w")

        # Analytics content (initially hidden)
        self.analytics_content = tk.Frame(self.analytics_frame, bg=COLORS['surface'])

        # Create main analytics container
        analytics_container = tk.Frame(self.analytics_content, bg=COLORS['surface'])
        analytics_container.pack(fill="both", expand=True, padx=10, pady=10)

        # Left side - Statistics cards
        stats_frame = tk.Frame(analytics_container, bg=COLORS['surface'])
        stats_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Statistics cards
        self._create_stats_cards(stats_frame)

        # Right side - Charts
        charts_frame = tk.Frame(analytics_container, bg=COLORS['surface'])
        charts_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))

        # Charts placeholder (will be filled when data is available)
        self.charts_container = tk.Frame(charts_frame, bg="white", relief="solid", bd=1)
        self.charts_container.pack(fill="both", expand=True, pady=10)

        # Placeholder text
        placeholder_label = tk.Label(
            self.charts_container,
            text="📈 Диаграммы появятся после поиска вакансий",
            font=("Arial", 12),
            bg="white",
            fg=COLORS['text_secondary']
        )
        placeholder_label.pack(expand=True)

        # Store reference for later use
        self.placeholder_label = placeholder_label

    def _create_stats_cards(self, parent_frame):
        """Create statistics cards."""
        # Create cards container
        cards_container = tk.Frame(parent_frame, bg=COLORS['surface'])
        cards_container.pack(fill="both", expand=True)

        # Statistics cards will be created/updated in _update_analytics method
        self.stats_cards_frame = cards_container

    def _update_analytics(self):
        """Update analytics data and charts."""
        try:
            vacancies = self.db_manager.get_all_vacancies()

            if not vacancies:
                # Show placeholder if no data
                self.placeholder_label.config(text="📈 Диаграммы появятся после поиска вакансий")
                return

            # Clear existing cards
            for widget in self.stats_cards_frame.winfo_children():
                widget.destroy()

            # Generate analytics data
            analytics_data = self._generate_analytics_data(vacancies)

            # Update statistics cards
            self._update_stats_cards(analytics_data)

            # Update charts
            self._update_charts(analytics_data)

        except Exception as e:
            logger.error(f"Error updating analytics: {e}")
            self.placeholder_label.config(text=f"Ошибка при загрузке аналитики: {e}")

    def _generate_analytics_data(self, vacancies):
        """Generate analytics data from vacancies."""
        # Salary analysis
        salaries = []
        salary_ranges = {'До 50k': 0, '50k-100k': 0, '100k-150k': 0, '150k-200k': 0, '200k+': 0}

        for vacancy in vacancies:
            if vacancy['salary_min'] or vacancy['salary_max']:
                # Use average of min/max for analysis
                avg_salary = (vacancy['salary_min'] or vacancy['salary_max'])
                salaries.append(avg_salary)

                # Categorize salary ranges
                if avg_salary < 50000:
                    salary_ranges['До 50k'] += 1
                elif avg_salary < 100000:
                    salary_ranges['50k-100k'] += 1
                elif avg_salary < 150000:
                    salary_ranges['100k-150k'] += 1
                elif avg_salary < 200000:
                    salary_ranges['150k-200k'] += 1
                else:
                    salary_ranges['200k+'] += 1

        # Location analysis
        locations = {}
        for vacancy in vacancies:
            location = vacancy['location'] or 'Не указан'
            locations[location] = locations.get(location, 0) + 1

        # Experience analysis
        experiences = {}
        for vacancy in vacancies:
            experience = vacancy['experience'] or 'Не указан'
            experiences[experience] = experiences.get(experience, 0) + 1

        # Remote work analysis
        remote_count = sum(1 for v in vacancies if v['remote'])
        office_count = len(vacancies) - remote_count

        # Calculate average salary
        avg_salary = sum(salaries) / len(salaries) if salaries else 0

        return {
            'total_vacancies': len(vacancies),
            'avg_salary': avg_salary,
            'salary_ranges': salary_ranges,
            'locations': locations,
            'experiences': experiences,
            'remote_vs_office': {'Удаленная': remote_count, 'Офис': office_count},
            'salaries': salaries
        }

    def _update_stats_cards(self, analytics_data):
        """Update statistics cards with data."""
        # Create cards grid
        cards_frame = tk.Frame(self.stats_cards_frame, bg=COLORS['surface'])
        cards_frame.pack(fill="both", expand=True)

        # Card 1: Total vacancies
        card1 = tk.Frame(cards_frame, bg="white", relief="solid", bd=1)
        card1.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        tk.Label(card1, text="📊 Всего вакансий", font=("Arial", 11, "bold"), bg=COLORS['primary'], fg="white").pack(fill="x")
        tk.Label(card1, text=str(analytics_data['total_vacancies']), font=("Arial", 24, "bold"), bg="white", fg=COLORS['primary']).pack(pady=10)

        # Card 2: Average salary
        card2 = tk.Frame(cards_frame, bg="white", relief="solid", bd=1)
        card2.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        tk.Label(card2, text="💰 Средняя зарплата", font=("Arial", 11, "bold"), bg=COLORS['secondary'], fg="white").pack(fill="x")
        avg_salary_text = f"{analytics_data['avg_salary']:,.0f} RUB" if analytics_data['avg_salary'] > 0 else "Не указана"
        tk.Label(card2, text=avg_salary_text, font=("Arial", 18, "bold"), bg="white", fg=COLORS['success']).pack(pady=10)

        # Card 3: Remote work
        card3 = tk.Frame(cards_frame, bg="white", relief="solid", bd=1)
        card3.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        tk.Label(card3, text="🏠 Удаленная работа", font=("Arial", 11, "bold"), bg=COLORS['accent'], fg="white").pack(fill="x")
        remote_percent = (analytics_data['remote_vs_office']['Удаленная'] / analytics_data['total_vacancies'] * 100) if analytics_data['total_vacancies'] > 0 else 0
        tk.Label(card3, text=f"{remote_percent:.1f}%", font=("Arial", 20, "bold"), bg="white", fg=COLORS['success']).pack(pady=10)

        # Card 4: With salary
        card4 = tk.Frame(cards_frame, bg="white", relief="solid", bd=1)
        card4.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
        tk.Label(card4, text="💵 С зарплатой", font=("Arial", 11, "bold"), bg=COLORS['primary'], fg="white").pack(fill="x")
        salary_percent = (len(analytics_data['salaries']) / analytics_data['total_vacancies'] * 100) if analytics_data['total_vacancies'] > 0 else 0
        tk.Label(card4, text=f"{salary_percent:.1f}%", font=("Arial", 20, "bold"), bg="white", fg=COLORS['secondary']).pack(pady=10)

        # Configure grid weights
        cards_frame.grid_columnconfigure(0, weight=1)
        cards_frame.grid_columnconfigure(1, weight=1)
        cards_frame.grid_rowconfigure(0, weight=1)
        cards_frame.grid_rowconfigure(1, weight=1)

    def _update_charts(self, analytics_data):
        """Update charts with analytics data."""
        try:
            # Clear existing charts
            for widget in self.charts_container.winfo_children():
                widget.destroy()

            # Import matplotlib here to avoid issues if not available
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib
            matplotlib.use('TkAgg')

            # Create figure with subplots
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(10, 8))
            fig.suptitle('Аналитика вакансий', fontsize=14, fontweight='bold')

            # Chart 1: Salary ranges pie chart
            salary_labels = list(analytics_data['salary_ranges'].keys())
            salary_values = list(analytics_data['salary_ranges'].values())
            if sum(salary_values) > 0:
                ax1.pie(salary_values, labels=salary_labels, autopct='%1.1f%%', startangle=90)
                ax1.set_title('Распределение зарплат', fontweight='bold')
            else:
                ax1.text(0.5, 0.5, 'Нет данных\nо зарплатах', ha='center', va='center', transform=ax1.transAxes)
                ax1.set_title('Распределение зарплат', fontweight='bold')

            # Chart 2: Remote vs Office pie chart
            remote_labels = list(analytics_data['remote_vs_office'].keys())
            remote_values = list(analytics_data['remote_vs_office'].values())
            if sum(remote_values) > 0:
                ax2.pie(remote_values, labels=remote_labels, autopct='%1.1f%%', startangle=90, colors=['#28A745', '#DC3545'])
                ax2.set_title('Удаленная vs Офисная работа', fontweight='bold')
            else:
                ax2.text(0.5, 0.5, 'Нет данных', ha='center', va='center', transform=ax2.transAxes)
                ax2.set_title('Удаленная vs Офисная работа', fontweight='bold')

            # Chart 3: Top locations bar chart
            location_labels = list(analytics_data['locations'].keys())[:5]  # Top 5
            location_values = list(analytics_data['locations'].values())[:5]
            if location_values:
                ax3.bar(range(len(location_labels)), location_values, color=COLORS['primary'])
                ax3.set_xticks(range(len(location_labels)))
                ax3.set_xticklabels(location_labels, rotation=45, ha='right')
                ax3.set_title('Топ локаций', fontweight='bold')
                ax3.set_ylabel('Количество вакансий')
            else:
                ax3.text(0.5, 0.5, 'Нет данных\nо локациях', ha='center', va='center', transform=ax3.transAxes)
                ax3.set_title('Топ локаций', fontweight='bold')

            # Chart 4: Experience distribution
            exp_labels = list(analytics_data['experiences'].keys())
            exp_values = list(analytics_data['experiences'].values())
            if exp_values:
                ax4.bar(range(len(exp_labels)), exp_values, color=COLORS['accent'])
                ax4.set_xticks(range(len(exp_labels)))
                ax4.set_xticklabels(exp_labels, rotation=45, ha='right')
                ax4.set_title('Распределение опыта работы', fontweight='bold')
                ax4.set_ylabel('Количество вакансий')
            else:
                ax4.text(0.5, 0.5, 'Нет данных\nоб опыте', ha='center', va='center', transform=ax4.transAxes)
                ax4.set_title('Распределение опыта работы', fontweight='bold')

            # Adjust layout
            plt.tight_layout()

            # Create canvas and embed in tkinter
            canvas = FigureCanvasTkAgg(fig, master=self.charts_container)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

        except ImportError:
            # If matplotlib is not available, show error message
            error_label = tk.Label(
                self.charts_container,
                text="❌ Matplotlib не установлен\nУстановите: pip install matplotlib",
                font=("Arial", 12),
                bg="white",
                fg=COLORS['error']
            )
            error_label.pack(expand=True)
        except Exception as e:
            error_label = tk.Label(
                self.charts_container,
                text=f"❌ Ошибка при создании диаграмм:\n{str(e)}",
                font=("Arial", 10),
                bg="white",
                fg=COLORS['error']
            )
            error_label.pack(expand=True)

    def _create_context_menu(self):
        """Create context menu for treeview."""
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Детали вакансии", command=self._show_vacancy_details)
        self.context_menu.add_command(label="Открыть в браузере", command=self._open_vacancy)
        self.context_menu.add_command(label="Удалить", command=self._delete_vacancy)

        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Double-1>", self._show_vacancy_details)

    def _show_context_menu(self, event):
        """Show context menu on right-click."""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def _create_status_bar(self):
        """Create status bar."""
        self.status_var = tk.StringVar()
        self.status_var.set("Готов к работе")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief="sunken", anchor="w")
        status_bar.pack(side="bottom", fill="x")

    def _start_search(self):
        """Start the search process in a separate thread."""
        keyword = self.keyword_var.get().strip()
        if not keyword:
            messagebox.showerror("Ошибка", "Введите ключевое слово для поиска")
            return

        # Disable search button and show progress
        self._set_search_state(False)
        self.progress_label.config(text="🚀 Идет поиск вакансий...")
        self.status_text.config(text="Подключение к hh.ru...")
        self.progress_var.set(10)
        self.status_var.set("Выполняется поиск вакансий...")

        # Start search in background thread
        search_thread = threading.Thread(target=self._perform_search,
                                       args=(keyword, self.location_var.get(), self.experience_var.get()))
        search_thread.daemon = True
        search_thread.start()

    def _perform_search(self, keyword: str, location: str, experience: str):
        """Perform the actual search operation."""
        try:
            # Clear old results first
            self.db_manager.clear_all_data()

            # Map user-friendly experience text to technical values
            experience_mapping = {
                'Нет опыта': 'noExperience',
                'От 1 до 3 лет': 'between1And3',
                'От 3 до 6 лет': 'between3And6',
                'Более 6 лет': 'moreThan6'
            }

            technical_experience = experience_mapping.get(experience, experience)

            # Update progress
            self.root.after(0, lambda: self._update_progress("🔍 Идет поиск вакансий...", "Подключение к hh.ru...", 20))

            # Parse vacancies
            vacancies = self.parser.search_vacancies(
                keyword=keyword,
                location=location if location else None,
                experience=technical_experience if technical_experience else None
            )

            if not vacancies:
                self.root.after(0, lambda: self._show_search_result("Вакансии не найдены", 0))
                return

            # Update progress
            self.root.after(0, lambda: self._update_progress("💾 Сохранение данных...", "Сохранение в базу данных...", 80))

            # Save to database
            saved_count = self.db_manager.save_vacancies_batch(vacancies)

            # Save search history
            self.db_manager.save_search_history(keyword, location, saved_count)

            # Update progress
            self.root.after(0, lambda: self._update_progress("✅ Поиск завершен!", f"Найдено вакансий: {saved_count}", 100))

            # Update UI
            self.root.after(0, lambda: self._show_search_result(f"Найдено и сохранено вакансий: {saved_count}", saved_count))

        except Exception as e:
            logger.error(f"Search error: {e}")
            self.root.after(0, lambda: self._show_search_error(str(e)))

    def _update_progress(self, label_text: str, status_text: str, progress_value: int):
        """Update progress bar and labels."""
        self.progress_label.config(text=label_text)
        self.status_text.config(text=status_text)
        self.progress_var.set(progress_value)

    def _show_search_result(self, message: str, count: int):
        """Show search completion message and update UI."""
        self.status_var.set(message)
        self._set_search_state(True)
        self._load_vacancies()

        if count > 0:
            messagebox.showinfo("Поиск завершен", message)

    def _show_search_error(self, error_message: str):
        """Show search error message."""
        self.status_var.set("Ошибка при выполнении поиска")
        self._set_search_state(True)
        messagebox.showerror("Ошибка поиска", f"Произошла ошибка:\n{error_message}")

    def _set_search_state(self, enabled: bool):
        """Enable or disable search controls."""
        state = "normal" if enabled else "disabled"

        # Find search button and disable/enable it
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.LabelFrame):
                for child in widget.winfo_children():
                    if isinstance(child, ttk.Button):
                        child.config(state=state)

    def _load_vacancies(self):
        """Load vacancies from database and display in treeview."""
        try:
            # Clear existing items
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Get vacancies from database
            vacancies = self.db_manager.get_all_vacancies()

            # Add to treeview - only title and company
            for vacancy in vacancies:
                self.tree.insert('', 'end', values=(
                    vacancy['title'],
                    vacancy['company']
                ))

            self.status_var.set(f"Загружено вакансий: {len(vacancies)}")

        except Exception as e:
            logger.error(f"Error loading vacancies: {e}")
            messagebox.showerror("Ошибка", f"Не удалось загрузить данные: {e}")

    def _show_vacancy_details(self, event=None):
        """Show detailed information about selected vacancy."""
        selection = self.tree.selection()
        if not selection:
            return

        item = selection[0]
        values = self.tree.item(item, 'values')

        # Find the corresponding vacancy in database
        vacancies = self.db_manager.get_all_vacancies()
        vacancy = None
        for v in vacancies:
            if v['title'] == values[0] and v['company'] == values[1]:
                vacancy = v
                break

        if not vacancy:
            return

        # Create detailed view window
        details_window = tk.Toplevel(self.root)
        details_window.title(f"📋 {vacancy['title']}")
        details_window.geometry("650x700")
        details_window.minsize(650, 700)  # Set minimum size
        details_window.configure(bg=COLORS['background'])

        # Main container frame
        container_frame = tk.Frame(details_window, bg=COLORS['surface'], relief="raised", bd=2)
        container_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Header frame with title
        header_frame = tk.Frame(container_frame, bg=COLORS['primary'], relief="flat")
        header_frame.pack(fill="x", padx=0, pady=0)

        title_label = tk.Label(
            header_frame,
            text=vacancy['title'],
            font=("Arial", 18, "bold"),
            bg=COLORS['primary'],
            fg="white",
            wraplength=600,
            justify="left"
        )
        title_label.pack(anchor="w", padx=20, pady=15)

        # Content frame
        content_frame = tk.Frame(container_frame, bg=COLORS['surface'], relief="flat")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Info cards frame
        info_frame = tk.Frame(content_frame, bg=COLORS['surface'])
        info_frame.pack(fill="both", expand=True)

        # Company card
        company_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        company_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            company_frame,
            text="🏢 Компания",
            font=("Arial", 11, "bold"),
            bg=COLORS['primary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            company_frame,
            text=vacancy['company'],
            font=("Arial", 12),
            bg="white",
            fg=COLORS['text'],
            wraplength=500,
            justify="left"
        ).pack(anchor="w", padx=10, pady=5)

        # Location card
        location_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        location_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            location_frame,
            text="📍 Местоположение",
            font=("Arial", 11, "bold"),
            bg=COLORS['primary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            location_frame,
            text=vacancy['location'] or "Не указано",
            font=("Arial", 12),
            bg="white",
            fg=COLORS['text'],
            wraplength=500,
            justify="left"
        ).pack(anchor="w", padx=10, pady=5)

        # Salary card
        salary_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        salary_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            salary_frame,
            text="💰 Зарплата",
            font=("Arial", 11, "bold"),
            bg=COLORS['secondary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)

        salary_text = ""
        if vacancy['salary_min'] or vacancy['salary_max']:
            if vacancy['salary_min'] == vacancy['salary_max']:
                salary_text = f"{vacancy['salary_min']}"
            else:
                salary_text = f"{vacancy['salary_min'] or ''}-{vacancy['salary_max'] or ''}"
            if vacancy['currency']:
                salary_text += f" {vacancy['currency']}"
        else:
            salary_text = "Не указана"

        tk.Label(
            salary_frame,
            text=salary_text,
            font=("Arial", 12, "bold"),
            bg="white",
            fg=COLORS['success'] if (vacancy['salary_min'] or vacancy['salary_max']) else COLORS['text_secondary'],
            wraplength=500,
            justify="left"
        ).pack(anchor="w", padx=10, pady=5)

        # Experience card
        experience_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        experience_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            experience_frame,
            text="👤 Опыт работы",
            font=("Arial", 11, "bold"),
            bg=COLORS['accent'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            experience_frame,
            text=vacancy['experience'] or "Не указан",
            font=("Arial", 12),
            bg="white",
            fg=COLORS['text'],
            wraplength=500,
            justify="left"
        ).pack(anchor="w", padx=10, pady=5)

        # Remote work card
        remote_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        remote_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            remote_frame,
            text="🏠 Удаленная работа",
            font=("Arial", 11, "bold"),
            bg=COLORS['primary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)

        remote_text = "Да" if vacancy['remote'] else "Нет"
        remote_color = COLORS['success'] if vacancy['remote'] else COLORS['error']

        tk.Label(
            remote_frame,
            text=remote_text,
            font=("Arial", 12, "bold"),
            bg="white",
            fg=remote_color,
            wraplength=500,
            justify="left"
        ).pack(anchor="w", padx=10, pady=5)

        # Link section
        link_frame = tk.Frame(content_frame, bg=COLORS['surface'])
        link_frame.pack(fill="x", pady=(10, 0))

        link_label = tk.Label(
            link_frame,
            text="🔗 Ссылка на вакансию:",
            font=("Arial", 12, "bold"),
            bg=COLORS['surface'],
            fg=COLORS['primary']
        )
        link_label.pack(anchor="w", pady=(0, 5))

        link_text = tk.Text(
            link_frame,
            height=2,
            width=60,
            font=("Arial", 10),
            bg=COLORS['background'],
            fg=COLORS['text'],
            wrap="word",
            relief="solid",
            bd=1
        )
        link_text.insert("1.0", vacancy['link'])
        link_text.config(state="disabled")
        link_text.pack(anchor="w", pady=(0, 10))

        # Buttons frame
        buttons_frame = tk.Frame(container_frame, bg=COLORS['surface'])
        buttons_frame.pack(fill="x", pady=(10, 0))

        # Open in browser button
        open_button = tk.Button(
            buttons_frame,
            text="🌐 Открыть в браузере",
            command=lambda: self._open_vacancy_url(vacancy['link']),
            bg=COLORS['secondary'],
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        open_button.pack(side="left", padx=(0, 10))

        # Close button
        close_button = tk.Button(
            buttons_frame,
            text="❌ Закрыть",
            command=details_window.destroy,
            bg=COLORS['accent'],
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        close_button.pack(side="right")

    def _open_vacancy_url(self, url: str):
        """Open vacancy URL in browser."""
        import webbrowser
        webbrowser.open(url)

    def _open_vacancy(self):
        """Open selected vacancy in browser."""
        selection = self.tree.selection()
        if not selection:
            return

        item = selection[0]
        values = self.tree.item(item, 'values')

        # Find the corresponding vacancy in database to get the link
        vacancies = self.db_manager.get_all_vacancies()
        for vacancy in vacancies:
            if (vacancy['title'] == values[0] and
                vacancy['company'] == values[1] and
                vacancy['location'] == values[2]):
                import webbrowser
                webbrowser.open(vacancy['link'])
                break

    def _delete_vacancy(self):
        """Delete selected vacancy from database."""
        selection = self.tree.selection()
        if not selection:
            return

        # Get the selected item values
        item = selection[0]
        values = self.tree.item(item, 'values')

        if messagebox.askyesno("Подтверждение", f"Удалить вакансию:\n{values[0]}?"):
            try:
                # Find the corresponding vacancy in database to get the ID
                vacancies = self.db_manager.get_all_vacancies()
                for vacancy in vacancies:
                    if (vacancy['title'] == values[0] and
                        vacancy['company'] == values[1] and
                        vacancy['location'] == values[2]):
                        # Delete from database
                        connection, cursor = self.db_manager._get_connection()
                        cursor.execute("DELETE FROM vacancies WHERE id = ?", (vacancy['id'],))
                        connection.commit()
                        break

                # Refresh the display
                self._load_vacancies()
                messagebox.showinfo("Успешно", "Вакансия удалена")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось удалить вакансию: {e}")

    def _export_to_excel(self):
        """Export vacancies to Excel file with formatting."""
        try:
            vacancies = self.db_manager.get_all_vacancies()
            if not vacancies:
                messagebox.showinfo("Информация", "Нет данных для экспорта")
                return

            # Select file location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if filename:
                # Create Excel file with formatting
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Вакансии"

                # Define headers
                headers = [
                    "Название вакансии",
                    "Компания",
                    "Местоположение",
                    "Зарплата",
                    "Опыт работы",
                    "Удаленная работа",
                    "Ссылка",
                    "Дата создания"
                ]

                # Add headers with styling
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="5E01A7", end_color="5E01A7", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Add data
                for row_num, vacancy in enumerate(vacancies, 2):
                    # Format salary
                    salary_text = ""
                    if vacancy['salary_min'] or vacancy['salary_max']:
                        if vacancy['salary_min'] == vacancy['salary_max']:
                            salary_text = f"{vacancy['salary_min']}"
                        else:
                            salary_text = f"{vacancy['salary_min'] or ''}-{vacancy['salary_max'] or ''}"
                        if vacancy['currency']:
                            salary_text += f" {vacancy['currency']}"
                    else:
                        salary_text = "Не указана"

                    # Format remote work
                    remote_text = "Да" if vacancy['remote'] else "Нет"

                    row_data = [
                        vacancy['title'],
                        vacancy['company'],
                        vacancy['location'] or "Не указано",
                        salary_text,
                        vacancy['experience'] or "Не указан",
                        remote_text,
                        vacancy['link'],
                        vacancy['created_at'][:19] if vacancy['created_at'] else ""
                    ]

                    for col_num, data in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=data)
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                # Auto-adjust column widths
                for col_num in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = 20

                # Make first column wider for titles
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['G'].width = 30  # Link column

                wb.save(filename)
                messagebox.showinfo("Успешно", f"Данные экспортированы в {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте: {e}")

    def _export_to_csv(self):
        """Export vacancies to CSV file with formatting."""
        try:
            vacancies = self.db_manager.get_all_vacancies()
            if not vacancies:
                messagebox.showinfo("Информация", "Нет данных для экспорта")
                return

            # Select file location
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filename:
                # Create CSV file with proper formatting
                import csv

                with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    # Define column headers
                    fieldnames = [
                        'Название вакансии',
                        'Компания',
                        'Местоположение',
                        'Зарплата',
                        'Опыт работы',
                        'Удаленная работа',
                        'Ссылка',
                        'Дата создания'
                    ]

                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                    # Write headers
                    writer.writeheader()

                    # Write data
                    for vacancy in vacancies:
                        # Format salary
                        salary_text = ""
                        if vacancy['salary_min'] or vacancy['salary_max']:
                            if vacancy['salary_min'] == vacancy['salary_max']:
                                salary_text = f"{vacancy['salary_min']}"
                            else:
                                salary_text = f"{vacancy['salary_min'] or ''}-{vacancy['salary_max'] or ''}"
                            if vacancy['currency']:
                                salary_text += f" {vacancy['currency']}"
                        else:
                            salary_text = "Не указана"

                        # Format remote work
                        remote_text = "Да" if vacancy['remote'] else "Нет"

                        writer.writerow({
                            'Название вакансии': vacancy['title'],
                            'Компания': vacancy['company'],
                            'Местоположение': vacancy['location'] or "Не указано",
                            'Зарплата': salary_text,
                            'Опыт работы': vacancy['experience'] or "Не указан",
                            'Удаленная работа': remote_text,
                            'Ссылка': vacancy['link'],
                            'Дата создания': vacancy['created_at'][:19] if vacancy['created_at'] else ""
                        })

                messagebox.showinfo("Успешно", f"Данные экспортированы в {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте: {e}")

    def _show_statistics(self):
        """Show database statistics with analytics."""
        try:
            stats = self.db_manager.get_statistics()
            vacancies = self.db_manager.get_all_vacancies()

            stats_window = tk.Toplevel(self.root)
            stats_window.title("📊 Статистика и аналитика")
            stats_window.geometry("600x700")
            stats_window.minsize(600, 700)
            stats_window.maxsize(600, 700)

            # Main container
            main_frame = tk.Frame(stats_window)
            main_frame.pack(fill="both", expand=True, padx=10, pady=10)

            # Title
            title_label = tk.Label(
                main_frame,
                text="📊 Статистика и аналитика вакансий",
                font=("Arial", 16, "bold"),
                fg=COLORS['primary']
            )
            title_label.pack(pady=(0, 20))

            # Basic statistics section
            basic_frame = tk.LabelFrame(main_frame, text="📋 Основная статистика", font=("Arial", 11, "bold"))
            basic_frame.pack(fill="x", pady=(0, 10))

            for key, value in stats.items():
                label_text = {
                    'total_vacancies': f'📊 Всего вакансий: {value}',
                    'vacancies_with_salary': f'💰 Вакансий с зарплатой: {value}',
                    'remote_vacancies': f'🏠 Удалённых вакансий: {value}',
                    'unique_companies': f'🏢 Уникальных компаний: {value}',
                    'unique_locations': f'📍 Уникальных локаций: {value}'
                }.get(key, f'{key}: {value}')

                ttk.Label(basic_frame, text=label_text, font=("Arial", 10)).pack(anchor="w", padx=10, pady=2)

            # Analytics section
            if vacancies:
                analytics_frame = tk.LabelFrame(main_frame, text="📈 Детальная аналитика", font=("Arial", 11, "bold"))
                analytics_frame.pack(fill="both", expand=True, pady=(0, 10))

                # Generate analytics data
                analytics_data = self._generate_analytics_data(vacancies)

                # Salary analysis
                salary_frame = tk.Frame(analytics_frame)
                salary_frame.pack(fill="x", pady=(0, 10))

                tk.Label(salary_frame, text="💰 Анализ зарплат:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)

                if analytics_data['avg_salary'] > 0:
                    tk.Label(salary_frame, text=f"📊 Средняя зарплата: {analytics_data['avg_salary']:,.0f} RUB", font=("Arial", 9)).pack(anchor="w", padx=20)
                else:
                    tk.Label(salary_frame, text="📊 Средняя зарплата: Не указана", font=("Arial", 9)).pack(anchor="w", padx=20)

                # Salary ranges
                tk.Label(salary_frame, text="📋 Распределение по диапазонам:", font=("Arial", 9, "bold")).pack(anchor="w", padx=20, pady=(5, 0))

                for range_name, count in analytics_data['salary_ranges'].items():
                    if count > 0:
                        tk.Label(salary_frame, text=f"  • {range_name}: {count}", font=("Arial", 8)).pack(anchor="w", padx=30)

                # Location analysis
                location_frame = tk.Frame(analytics_frame)
                location_frame.pack(fill="x", pady=(0, 10))

                tk.Label(location_frame, text="📍 Топ локаций:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)

                # Show top 5 locations
                sorted_locations = sorted(analytics_data['locations'].items(), key=lambda x: x[1], reverse=True)[:5]
                for location, count in sorted_locations:
                    tk.Label(location_frame, text=f"  • {location}: {count}", font=("Arial", 9)).pack(anchor="w", padx=20)

                # Experience analysis
                exp_frame = tk.Frame(analytics_frame)
                exp_frame.pack(fill="x", pady=(0, 10))

                tk.Label(exp_frame, text="👤 Распределение по опыту:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)

                for experience, count in analytics_data['experiences'].items():
                    if count > 0:
                        tk.Label(exp_frame, text=f"  • {experience}: {count}", font=("Arial", 9)).pack(anchor="w", padx=20)

                # Remote work analysis
                remote_frame = tk.Frame(analytics_frame)
                remote_frame.pack(fill="x", pady=(0, 10))

                tk.Label(remote_frame, text="🏠 Удаленная работа:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)

                remote_percent = (analytics_data['remote_vs_office']['Удаленная'] / analytics_data['total_vacancies'] * 100) if analytics_data['total_vacancies'] > 0 else 0
                salary_percent = (len(analytics_data['salaries']) / analytics_data['total_vacancies'] * 100) if analytics_data['total_vacancies'] > 0 else 0

                tk.Label(remote_frame, text=f"📊 Удаленная работа: {remote_percent:.1f}%", font=("Arial", 9)).pack(anchor="w", padx=20)
                tk.Label(remote_frame, text=f"💵 С указанной зарплатой: {salary_percent:.1f}%", font=("Arial", 9)).pack(anchor="w", padx=20)
            else:
                no_data_frame = tk.Frame(main_frame)
                no_data_frame.pack(fill="both", expand=True)

                tk.Label(
                    no_data_frame,
                    text="📈 Для аналитики нужно выполнить поиск вакансий",
                    font=("Arial", 12),
                    fg=COLORS['text_secondary']
                ).pack(expand=True)

            # Close button
            close_button = tk.Button(
                main_frame,
                text="❌ Закрыть",
                command=stats_window.destroy,
                bg=COLORS['accent'],
                fg="white",
                font=("Arial", 10, "bold"),
                relief="flat",
                padx=20,
                pady=8
            )
            close_button.pack(pady=20)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось получить статистику: {e}")

    def _clear_database(self):
        """Clear all data from database."""
        if messagebox.askyesno("Подтверждение",
                             "Вы уверены, что хотите удалить ВСЕ данные из базы данных?\n\n"
                             "Это действие нельзя отменить!"):
            try:
                if self.db_manager.clear_all_data():
                    messagebox.showinfo("Успешно", "База данных очищена")
                    # Force refresh the display
                    self._load_vacancies()
                    # Update status bar
                    self.status_var.set("База данных очищена")
                else:
                    messagebox.showerror("Ошибка", "Не удалось очистить базу данных")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при очистке базы данных: {e}")

    def _show_about(self):
        """Show about dialog with modern styling."""
        about_window = tk.Toplevel(self.root)
        about_window.title("ℹ️ О программе")
        about_window.geometry("550x750")
        about_window.minsize(550, 750)
        about_window.maxsize(550, 750)
        about_window.configure(bg=COLORS['background'])

        # Main container
        main_frame = tk.Frame(about_window, bg=COLORS['surface'], relief="raised", bd=2)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Header with icon
        header_frame = tk.Frame(main_frame, bg=COLORS['primary'], relief="flat")
        header_frame.pack(fill="x", padx=0, pady=0)

        title_label = tk.Label(
            header_frame,
            text="MarketScope",
            font=("Arial", 20, "bold"),
            bg=COLORS['primary'],
            fg="white"
        )
        title_label.pack(anchor="w", padx=20, pady=15)

        # Content frame
        content_frame = tk.Frame(main_frame, bg=COLORS['surface'], relief="flat")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # App info
        info_frame = tk.Frame(content_frame, bg=COLORS['surface'])
        info_frame.pack(fill="both", expand=True)

        # Version info
        version_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        version_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            version_frame,
            text="📦 Версия программы",
            font=("Arial", 11, "bold"),
            bg=COLORS['primary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            version_frame,
            text="HH.ru Vacancy Scraper v1.0",
            font=("Arial", 12, "bold"),
            bg="white",
            fg=COLORS['primary']
        ).pack(anchor="w", padx=10, pady=5)

        # Description
        desc_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        desc_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            desc_frame,
            text="📋 Описание",
            font=("Arial", 11, "bold"),
            bg=COLORS['secondary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            desc_frame,
            text="Программа для автоматического сбора и анализа вакансий с сайта hh.ru.\n\n"
                 "Возможности:\n"
                 "• Поиск вакансий по ключевым словам\n"
                 "• Фильтрация по локации и опыту работы\n"
                 "• Детальный анализ результатов\n"
                 "• Экспорт данных в Excel и CSV\n"
                 "• Статистика и аналитика",
            font=("Arial", 10),
            bg="white",
            fg=COLORS['text'],
            justify="left",
            wraplength=400
        ).pack(anchor="w", padx=10, pady=5)

        # Purpose
        purpose_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        purpose_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            purpose_frame,
            text="🎯 Назначение",
            font=("Arial", 11, "bold"),
            bg=COLORS['accent'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)
        tk.Label(
            purpose_frame,
            text="Разработано для автоматизации поиска работы и анализа рынка труда.\n"
                 "Помогает быстро находить релевантные вакансии и анализировать их.",
            font=("Arial", 10),
            bg="white",
            fg=COLORS['text'],
            justify="left",
            wraplength=400
        ).pack(anchor="w", padx=10, pady=5)

        # Developer info
        dev_frame = tk.Frame(info_frame, bg="white", relief="solid", bd=1)
        dev_frame.pack(fill="x", pady=(0, 10), padx=0)
        tk.Label(
            dev_frame,
            text="👨‍💻 Разработчик",
            font=("Arial", 11, "bold"),
            bg=COLORS['primary'],
            fg="white"
        ).pack(fill="x", padx=10, pady=5)

        # Try to load developer logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'skrauch.png')
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                logo_image = logo_image.resize((104, 47), Image.Resampling.LANCZOS)
                self.dev_logo_photo = ImageTk.PhotoImage(logo_image)

                logo_label = tk.Label(dev_frame, image=self.dev_logo_photo, bg="white")
                logo_label.pack(anchor="center", pady=(5, 10))
        except Exception as e:
            logger.warning(f"Could not load developer logo: {e}")

        tk.Label(
            dev_frame,
            text="© 2025 Все права защищены",
            font=("Arial", 10, "bold"),
            bg="white",
            fg=COLORS['primary'],
            justify="center"
        ).pack(anchor="center", padx=10, pady=5)

        # Buttons frame
        buttons_frame = tk.Frame(main_frame, bg=COLORS['surface'])
        buttons_frame.pack(fill="x", pady=(10, 0))

        # Close button
        close_button = tk.Button(
            buttons_frame,
            text="❌ Закрыть",
            command=about_window.destroy,
            bg=COLORS['accent'],
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        close_button.pack(side="right")


def create_app():
    """Create and return the main application window."""
    root = tk.Tk()
    app = VacancyApp(root)
    return root


def run_app():
    """Run the main application."""
    root = create_app()
    root.mainloop()


if __name__ == "__main__":
    run_app()
